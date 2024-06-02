# from selenium import webdriver
from openpyxl import load_workbook

class LibExcel:
    @staticmethod
    def GetIsRunData(filePath, sheetName):
        wb = load_workbook(filePath)
        ws = wb[sheetName]

        rowsWithRun = []

        for row in range(2, ws.max_row + 1):
            if ws.cell(row=row, column=1).value and \
                    ws.cell(row=row, column=1).value.strip().upper() == "RUN".upper():
                rowsWithRun.append(row)

        return rowsWithRun

    @staticmethod
    def getDataExcel(filePath, columnName, sheetName):
        wb = load_workbook(filePath)
        ws = wb[sheetName]

        rowHasRun = 0
        for row in range(2, ws.max_row + 1):
            if ws.cell(row=row, column=1).value and \
                    ws.cell(row=row, column=1).value.strip().upper() == "RUN".upper():
                rowHasRun = row
                break

        column_index = -1
        for col in range(1, ws.max_column + 1):
            if ws.cell(row=1, column=col).value == columnName:
                column_index = col
                break

        if column_index != -1:
            return ws.cell(row=rowHasRun, column=column_index).value
        else:
            return None  # Column tidak ditemukan
        